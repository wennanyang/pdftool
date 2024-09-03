import shutil
import traceback
import re
from pathlib import Path
import PyPDF2 as PyPDF
from openpyxl import load_workbook, Workbook

LER_PATTERN = re.compile(r"(?<=Luminaire Efficacy Rating \(LER\): )\d+(\.\d{1,})?")
TRLL_PATTERN = re.compile(r"(?<=Total Rated Lamp Lumens: )\d+(\.\d{1,})?(?= lm)")
def make_PDF_result(pdf_dir : Path, exp_dir_name=Path("无法提取的PDF"),
                     save_name=Path("提取结果.xlsx"),progress_callback=None):
    title=["文件路径","Luminaire Efficacy Rating","Total Rated Lamp Lumens"]
    ## 把放线异常的项目copy的文件夹和对应的列表
    if exp_dir_name.exists():
        shutil.rmtree(exp_dir_name)
    exp_dir_name.mkdir(parents=True, exist_ok=True)
    ## 输出的excel
    if save_name.exists():
        save_name.unlink
    wb = Workbook()
    ws = wb.active
    ws.title = "提取结果"
    wb.save(save_name)
    ## 初始化
    wb = load_workbook(save_name)
    ws = wb["提取结果"]
    ws.append(title)
    # 提取失败的文件
    fail_list = []
    exp_count = 1
    pdf_files = list(pdf_dir.glob("*"))
    for i, pdf_file in enumerate(pdf_files):
        print(i)
        if progress_callback is not None:
            progress_callback(value=(i + 1) / len(pdf_files) * 100, 
                              description=f"正在提取属性:{pdf_file.name}")
        try:
            result = get_PDF_result(pdf_file)
            if result is None:
                fail_list.append(pdf_file.resolve())
                continue
        except Exception as e:
                ## 写入日志
            with open("log", 'a', encoding='utf-8') as f:
                f.write(f"{exp_count}'\t'{pdf_file.resolve()}'\n'{traceback.format_exc()}'\n'")
                exp_count += 1
            fail_list.append(pdf_file.resolve())
            continue
        ws.append(result)
    wb.save(save_name)
    for file in fail_list:   
        try:
            shutil.copy(file, exp_dir_name.joinpath(file.name))
        except Exception as e:
            print(e)
            continue
    

    
def get_PDF_result(pdf_path : Path):
    pdf = PyPDF.PdfReader(pdf_path)
    # label = pdf.pq('LTTextLineHorizontal:contains("Your Label")')
    text = pdf.pages[0].extract_text()
    LER_match = LER_PATTERN.search(text)
    TRLL_match = TRLL_PATTERN.search(text)
    if LER_match is not None and TRLL_match is not None:
        return pdf_path.resolve().as_posix(), LER_match.group(), TRLL_match.group()
    return None

def main(pdf_dir=None, progress_callback=None, args=None):
    if pdf_dir is not None :
        make_PDF_result(pdf_dir=Path(pdf_dir), 
                        progress_callback=progress_callback)

if __name__ == '__main__':
    # path = Path(r"E:\Code\测试\WPL01-25W-B28D5857-BFA1-FP.pdf")

    # txt = get_PDF_result(path)
    # print(txt)
    main(pdf_dir = "E:\Code\测试")
    # Total Rated Lamp Lumens
    # Luminaire Efficacy Rating (LER)
    # print("1.成果汇总表（厂房(自编号初期雨水及事故水池)）" == "1.成果汇总表（厂房(自编号初期雨水及事故水池)）")